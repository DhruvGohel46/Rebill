import os
import re
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import Config


class ExcelReportBuilder:
    """
    Standardized helper class for generating branded, publication-grade InfoOS Excel (.xlsx) reports.
    Implements unified header blocks, column headers, totals-row styling, number formatting,
    zebra striping, and auto-fitting column widths.
    """

    # Brand Colors
    COLOR_BRAND_ORANGE = "F97316"
    COLOR_DARK_HEADER = "1E2025"
    COLOR_TOTALS_FILL = "FFF3E8"
    COLOR_ZEBRA_ODD = "FFFFFF"
    COLOR_ZEBRA_EVEN = "FAFAFA"
    COLOR_METRIC_BG = "F3F4F6"
    COLOR_BORDER = "E5E7EB"
    COLOR_BORDER_DARK = "CBD5E1"
    COLOR_MUTED_TEXT = "6B7280"

    # Number Formats
    FMT_CURRENCY = "₹#,##0.00"
    FMT_INTEGER = "#,##0"
    FMT_DECIMAL = "#,##0.00"
    FMT_PERCENT = "0.0%"
    FMT_DATE = "DD-MMM-YYYY"
    FMT_DATETIME = "DD-MMM-YYYY HH:MM"

    def __init__(self):
        self._load_shop_profile()

    def _load_shop_profile(self):
        """Fetch shop name, address, and GST/Tax ID from database settings or config."""
        self.shop_name = Config.SHOP_NAME or "InfoOS Store"
        self.shop_address = ""
        self.gst_number = ""
        try:
            from services.db_service import DatabaseService

            db_service = DatabaseService()
            settings = db_service.get_all_settings()
            if settings.get("shop_name") and str(settings.get("shop_name")).strip():
                self.shop_name = str(settings.get("shop_name")).strip()
            if settings.get("shop_address") and str(settings.get("shop_address")).strip():
                self.shop_address = str(settings.get("shop_address")).strip()
            gst_val = settings.get("gst_number") or settings.get("gst_no") or settings.get("tax_id")
            if gst_val and str(gst_val).strip():
                self.gst_number = str(gst_val).strip()
        except Exception:
            pass

    def create_workbook(self) -> openpyxl.Workbook:
        """Create a fresh openpyxl workbook."""
        self._load_shop_profile()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default blank sheet so named sheets are first
        return wb

    def get_safe_filename(self, report_type: str, date_or_range: str) -> str:
        """
        Generate standardized filename: InfoOS_{ReportType}_{DateOrRange}_{ShopName}.xlsx
        Guarantees safe alphanumeric characters and no spaces.
        """
        self._load_shop_profile()
        safe_report = re.sub(r"[^a-zA-Z0-9]", "", report_type)
        safe_date = re.sub(r"[^a-zA-Z0-9\-_]", "", str(date_or_range))
        safe_shop = re.sub(r"[^a-zA-Z0-9]", "", self.shop_name) or "Store"
        return f"InfoOS_{safe_report}_{safe_date}_{safe_shop}.xlsx"

    def write_branded_header(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        report_title: str,
        date_range_label: str,
        num_columns: int = 6,
    ):
        """
        Write standard 5-row branded header block merged across used columns:
        Row 1: Shop Name (16pt bold)
        Row 2: Address + GST/Tax ID (10pt grey)
        Row 3: Report Title (13pt bold orange #F97316)
        Row 4: Date Range + Generated on (9pt grey, right-aligned)
        Row 5: Blank spacer
        """
        self._load_shop_profile()
        end_col_letter = get_column_letter(max(num_columns, 4))

        # Row 1: Shop Name
        ws.merge_cells(f"A1:{end_col_letter}1")
        cell_r1 = ws["A1"]
        cell_r1.value = self.shop_name
        cell_r1.font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        cell_r1.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        # Row 2: Address + Tax ID
        ws.merge_cells(f"A2:{end_col_letter}2")
        cell_r2 = ws["A2"]
        sub_info_parts = []
        if self.shop_address:
            sub_info_parts.append(self.shop_address)
        if self.gst_number:
            sub_info_parts.append(f"GST/Tax: {self.gst_number}")
        cell_r2.value = (
            " | ".join(sub_info_parts) if sub_info_parts else "Point of Sale & Business Operations"
        )
        cell_r2.font = Font(name="Calibri", size=10, color=self.COLOR_MUTED_TEXT)
        cell_r2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16

        # Row 3: Report Title
        ws.merge_cells(f"A3:{end_col_letter}3")
        cell_r3 = ws["A3"]
        cell_r3.value = report_title.upper()
        cell_r3.font = Font(name="Calibri", size=13, bold=True, color=self.COLOR_BRAND_ORANGE)
        cell_r3.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[3].height = 20

        # Row 4: Range & Generated Timestamp
        ws.merge_cells(f"A4:{end_col_letter}4")
        cell_r4 = ws["A4"]
        now_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        cell_r4.value = f"Period: {date_range_label}    |    Generated on: {now_str}"
        cell_r4.font = Font(name="Calibri", size=9, italic=True, color=self.COLOR_MUTED_TEXT)
        cell_r4.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[4].height = 16

        # Row 5: Blank spacer
        ws.row_dimensions[5].height = 10

    def write_metric_cards(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        start_row: int,
        metrics: List[Dict[str, Any]],
        cols_per_card: int = 2,
    ) -> int:
        """
        Write a grid of polished executive KPI metric cards.
        Each metric is: {"label": "Total Sales", "value": 4500.0, "format": "currency"|"number"|"text"}
        Returns the next available row.
        """
        thin_border = Border(
            left=Side(style="thin", color=self.COLOR_BORDER_DARK),
            right=Side(style="thin", color=self.COLOR_BORDER_DARK),
            top=Side(style="thin", color=self.COLOR_BORDER_DARK),
            bottom=Side(style="thin", color=self.COLOR_BORDER_DARK),
        )

        current_col = 1
        current_row = start_row

        ws.row_dimensions[current_row].height = 16
        ws.row_dimensions[current_row + 1].height = 24

        for item in metrics:
            col_start = current_col
            col_end = current_col + cols_per_card - 1
            col_start_letter = get_column_letter(col_start)
            col_end_letter = get_column_letter(col_end)

            # Merge Label cells
            ws.merge_cells(f"{col_start_letter}{current_row}:{col_end_letter}{current_row}")
            label_cell = ws[f"{col_start_letter}{current_row}"]
            label_cell.value = str(item.get("label", "")).upper()
            label_cell.font = Font(name="Calibri", size=8.5, bold=True, color=self.COLOR_MUTED_TEXT)
            label_cell.fill = PatternFill(
                start_color=self.COLOR_METRIC_BG, end_color=self.COLOR_METRIC_BG, fill_type="solid"
            )
            label_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Merge Value cells
            ws.merge_cells(f"{col_start_letter}{current_row+1}:{col_end_letter}{current_row+1}")
            val_cell = ws[f"{col_start_letter}{current_row+1}"]
            val = item.get("value", "")
            val_format = item.get("format", "text")

            val_cell.value = val
            val_cell.font = Font(name="Calibri", size=13, bold=True, color="0F172A")
            val_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            val_cell.alignment = Alignment(horizontal="center", vertical="center")

            if val_format == "currency":
                val_cell.number_format = self.FMT_CURRENCY
            elif val_format == "percent":
                val_cell.number_format = self.FMT_PERCENT
            elif val_format == "number":
                val_cell.number_format = self.FMT_INTEGER

            # Apply border to all bounding cells in card
            for r in range(current_row, current_row + 2):
                for c in range(col_start, col_end + 1):
                    ws.cell(row=r, column=c).border = thin_border

            current_col += cols_per_card

        # Return row after metric cards + spacer
        ws.row_dimensions[current_row + 2].height = 12
        return current_row + 3

    def write_table(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        start_row: int,
        headers: List[str],
        data_rows: List[List[Any]],
        col_formats: Optional[List[str]] = None,
        col_alignments: Optional[List[str]] = None,
        totals_row: Optional[List[Any]] = None,
        section_title: Optional[str] = None,
        freeze_header: bool = True,
    ) -> int:
        """
        Write a publication-quality data table with:
        - Optional section header
        - Frozen dark column header row (#1E2025, bold white text)
        - Zebra-striped data rows
        - Configured column number formats & alignments
        - Styled totals row (#FFF3E8 fill, top/bottom border)
        Returns next available row index.
        """
        current_row = start_row

        if section_title:
            ws.cell(row=current_row, column=1, value=section_title).font = Font(
                name="Calibri", size=11, bold=True, color=self.COLOR_BRAND_ORANGE
            )
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        header_row_idx = current_row

        # Freeze panes at column headers if requested (only once per sheet)
        if freeze_header and not ws.freeze_panes:
            ws.freeze_panes = f"A{header_row_idx + 1}"

        # Write Column Headers
        ws.row_dimensions[header_row_idx].height = 26
        header_fill = PatternFill(
            start_color=self.COLOR_DARK_HEADER, end_color=self.COLOR_DARK_HEADER, fill_type="solid"
        )
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_border = Border(
            left=Side(style="thin", color="374151"),
            right=Side(style="thin", color="374151"),
            top=Side(style="medium", color="111827"),
            bottom=Side(style="medium", color="111827"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            align = (
                col_alignments[col_idx - 1]
                if col_alignments and len(col_alignments) >= col_idx
                else "left"
            )
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = header_border

        current_row += 1

        # Write Data Rows (Zebra striped)
        thin_border = Border(
            left=Side(style="thin", color=self.COLOR_BORDER),
            right=Side(style="thin", color=self.COLOR_BORDER),
            top=Side(style="thin", color=self.COLOR_BORDER),
            bottom=Side(style="thin", color=self.COLOR_BORDER),
        )

        fill_odd = PatternFill(
            start_color=self.COLOR_ZEBRA_ODD, end_color=self.COLOR_ZEBRA_ODD, fill_type="solid"
        )
        fill_even = PatternFill(
            start_color=self.COLOR_ZEBRA_EVEN, end_color=self.COLOR_ZEBRA_EVEN, fill_type="solid"
        )
        data_font = Font(name="Calibri", size=10, color="1F2937")

        for row_idx, row_data in enumerate(data_rows):
            ws.row_dimensions[current_row].height = 20
            row_fill = fill_even if row_idx % 2 == 1 else fill_odd

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border

                align = (
                    col_alignments[col_idx - 1]
                    if col_alignments and len(col_alignments) >= col_idx
                    else "left"
                )
                cell.alignment = Alignment(horizontal=align, vertical="center")

                # Apply Number Format
                fmt = (
                    col_formats[col_idx - 1]
                    if col_formats and len(col_formats) >= col_idx
                    else None
                )
                if fmt:
                    cell.number_format = fmt

            current_row += 1

        # Write Totals Row if provided
        if totals_row:
            ws.row_dimensions[current_row].height = 24
            totals_fill = PatternFill(
                start_color=self.COLOR_TOTALS_FILL,
                end_color=self.COLOR_TOTALS_FILL,
                fill_type="solid",
            )
            totals_font = Font(name="Calibri", size=10.5, bold=True, color="0F172A")
            totals_border = Border(
                left=Side(style="thin", color=self.COLOR_BORDER_DARK),
                right=Side(style="thin", color=self.COLOR_BORDER_DARK),
                top=Side(style="medium", color=self.COLOR_BRAND_ORANGE),
                bottom=Side(style="double", color="1E2025"),
            )

            for col_idx, val in enumerate(totals_row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = totals_font
                cell.fill = totals_fill
                cell.border = totals_border

                align = (
                    col_alignments[col_idx - 1]
                    if col_alignments and len(col_alignments) >= col_idx
                    else "left"
                )
                cell.alignment = Alignment(horizontal=align, vertical="center")

                fmt = (
                    col_formats[col_idx - 1]
                    if col_formats and len(col_formats) >= col_idx
                    else None
                )
                if fmt:
                    cell.number_format = fmt

            current_row += 1

        # Spacer row after table
        ws.row_dimensions[current_row].height = 14
        return current_row + 1

    def autofit_column_widths(
        self, ws: openpyxl.worksheet.worksheet.Worksheet, min_width: int = 12, max_width: int = 48
    ):
        """
        Auto-fit column widths according to content length with a safe minimum.
        Handles merged cells and multi-line strings gracefully.
        """
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = min_width
            for cell in col:
                if cell.row in [1, 2, 3, 4]:  # Skip full-width merged header rows
                    continue
                if cell.value is not None:
                    # Treat newline formatted values
                    lines = str(cell.value).split("\n")
                    for l in lines:
                        if len(l) > max_len:
                            max_len = len(l)
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

    def write_empty_state_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        report_title: str,
        date_range_label: str,
        message: str = "No transactions or records found for the selected period.",
    ):
        """Standardized empty state for non-operating dates/periods."""
        self.write_branded_header(ws, report_title, date_range_label, num_columns=5)
        ws.merge_cells("A7:E8")
        cell = ws["A7"]
        cell.value = f"ℹ️  {message}"
        cell.font = Font(name="Calibri", size=11, bold=True, color=self.COLOR_MUTED_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        self.autofit_column_widths(ws)
