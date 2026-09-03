import json
import os
import openpyxl
import pytest
from datetime import date, datetime
from models import db, ItemGroup, Category, Product, Bill, Expense
from services.excel_xlsx_service import ExcelXLSXService
from services.excel_service import ExcelService
from services.summary_service import SummaryService
from services.db_service import DatabaseService


@pytest.fixture
def setup_catalog_and_bills(app):
    with app.app_context():
        # Clean existing test data
        Bill.query.delete()
        Expense.query.delete()
        Product.query.delete()
        Category.query.delete()
        ItemGroup.query.delete()
        db.session.commit()

        # Set shop settings via DatabaseService
        db_svc = DatabaseService()
        db_svc.update_settings_bulk(
            [
                {"key": "shop_name", "value": "Spice Kingdom"},
                {"key": "shop_address", "value": "42 Curry Lane, Food City"},
                {"key": "gst_number", "value": "GSTIN27AABCS1234F1Z5"},
            ]
        )

        # Create ItemGroups
        grp_bev = ItemGroup(name="Beverages", color="#3B82F6", icon="coffee")
        grp_food = ItemGroup(name="Main Kitchen", color="#10B981", icon="utensils")
        db.session.add_all([grp_bev, grp_food])
        db.session.flush()

        # Create Categories
        cat_tea = Category(name="Tea & Coffee", group_id=grp_bev.id)
        cat_shakes = Category(name="Milkshakes", group_id=grp_bev.id)
        cat_burgers = Category(name="Burgers", group_id=grp_food.id)
        db.session.add_all([cat_tea, cat_shakes, cat_burgers])
        db.session.flush()

        # Create Products
        p1 = Product(
            product_id="P101",
            name="Masala Chai",
            price=20.0,
            category_id=cat_tea.id,
            category="Tea & Coffee",
        )
        p2 = Product(
            product_id="P102",
            name="Mango Shake",
            price=60.0,
            category_id=cat_shakes.id,
            category="Milkshakes",
        )
        p3 = Product(
            product_id="P103",
            name="Veg Cheese Burger",
            price=120.0,
            category_id=cat_burgers.id,
            category="Burgers",
        )
        db.session.add_all([p1, p2, p3])
        db.session.commit()

        # Create today's bills
        today_dt = datetime.now()
        bill1 = Bill(
            bill_no=1001,
            total_amount=160.0,
            payment_method="UPI",
            status="PAID",
            created_at=today_dt,
            items=json.dumps(
                [
                    {"product_id": "P101", "name": "Masala Chai", "quantity": 2, "price": 20.0, "subtotal": 40.0, "cost_price": 8.0, "category": "Tea & Coffee"},
                    {"product_id": "P103", "name": "Veg Cheese Burger", "quantity": 1, "price": 120.0, "subtotal": 120.0, "cost_price": 45.0, "category": "Burgers"},
                ]
            ),
        )
        bill2 = Bill(
            bill_no=1002,
            total_amount=120.0,
            payment_method="CASH",
            status="CONFIRMED",
            created_at=today_dt,
            items=json.dumps(
                [
                    {"product_id": "P102", "name": "Mango Shake", "quantity": 2, "price": 60.0, "subtotal": 120.0, "cost_price": 25.0, "category": "Milkshakes"},
                ]
            ),
        )
        db.session.add_all([bill1, bill2])
        db.session.commit()

        yield {
            "group_bev": grp_bev,
            "group_food": grp_food,
            "cat_tea": cat_tea,
            "cat_shakes": cat_shakes,
            "cat_burgers": cat_burgers,
            "products": [p1, p2, p3],
            "bills": [bill1, bill2],
        }


def test_dynamic_shop_name_in_reports(app, setup_catalog_and_bills):
    """Test that report generator immediately reflects dynamic shop name from settings."""
    with app.app_context():
        xlsx_svc = ExcelXLSXService()
        filepath = xlsx_svc.export_daily_sales_report(date.today().strftime("%Y-%m-%d"))

        # Verify filename contains dynamic shop name
        assert "SpiceKingdom" in os.path.basename(filepath)
        assert os.path.exists(filepath)

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Summary"]
        # Branded header shop name is at row 1, col 1 (A1)
        header_title = ws.cell(row=1, column=1).value
        assert "SPICE KINGDOM" in header_title.upper()

        # Row 2 (A2) contains address and GST
        sub_info = ws.cell(row=2, column=1).value
        assert "42 Curry Lane" in sub_info
        assert "GSTIN27AABCS1234F1Z5" in sub_info

        # Dynamically change shop name in settings
        db_svc = DatabaseService()
        db_svc.update_settings_bulk([{"key": "shop_name", "value": "Royal Grand Cafe"}])

        # Generate report again
        filepath2 = xlsx_svc.export_daily_sales_report(date.today().strftime("%Y-%m-%d"))
        assert "RoyalGrandCafe" in os.path.basename(filepath2)
        wb2 = openpyxl.load_workbook(filepath2)
        ws2 = wb2["Summary"]
        assert "ROYAL GRAND CAFE" in ws2.cell(row=1, column=1).value.upper()


def test_daily_sales_group_and_category_breakdown(app, setup_catalog_and_bills):
    """Test Daily Sales Report has Group & Category summary table on Sheet 1 and Group column on Sheet 2."""
    with app.app_context():
        xlsx_svc = ExcelXLSXService()
        filepath = xlsx_svc.export_daily_sales_report(date.today().strftime("%Y-%m-%d"))

        wb = openpyxl.load_workbook(filepath)
        assert "Summary" in wb.sheetnames
        assert "Item-Wise Breakdown" in wb.sheetnames

        ws_sum = wb["Summary"]
        # Find Group & Category table section
        found_gc_section = False
        for r in range(1, 40):
            val = ws_sum.cell(row=r, column=1).value
            if val and "Group & Category" in str(val):
                found_gc_section = True
                break
        assert found_gc_section, "Group & Category section title not found in Summary sheet"

        ws_items = wb["Item-Wise Breakdown"]
        headers = [ws_items.cell(row=6, column=c).value for c in range(1, 9)]
        assert "Group" in headers
        assert "Category" in headers
        assert "Product Name" in headers

        # Verify row content has resolved group and category
        row7 = [ws_items.cell(row=7, column=c).value for c in range(1, 9)]
        assert row7[0] in ["Beverages", "Main Kitchen"]  # Group column


def test_weekly_sales_group_and_category_breakdown(app, setup_catalog_and_bills):
    """Test Weekly Sales Summary has Group & Category table and dedicated sheet."""
    with app.app_context():
        xlsx_svc = ExcelXLSXService()
        filepath = xlsx_svc.export_weekly_sales_summary(date.today().strftime("%Y-%m-%d"))

        wb = openpyxl.load_workbook(filepath)
        assert "Week Overview" in wb.sheetnames
        assert "Group & Category Breakdown" in wb.sheetnames
        assert "Product Performance" in wb.sheetnames

        ws_grp = wb["Group & Category Breakdown"]
        headers = [ws_grp.cell(row=6, column=c).value for c in range(1, 7)]
        assert "Group" in headers
        assert "Category" in headers
        assert "Units Sold (Week)" in headers
        assert "Revenue (Week)" in headers

        ws_prod = wb["Product Performance"]
        prod_headers = [ws_prod.cell(row=6, column=c).value for c in range(1, 8)]
        assert "Group" in prod_headers
        assert "Category" in prod_headers
        assert "Product" in prod_headers


def test_monthly_sales_group_and_category_breakdown(app, setup_catalog_and_bills):
    """Test Monthly Sales Summary has dedicated Group & Category Breakdown sheet."""
    with app.app_context():
        xlsx_svc = ExcelXLSXService()
        today = date.today()
        filepath = xlsx_svc.export_monthly_sales_summary(today.month, today.year)

        wb = openpyxl.load_workbook(filepath)
        assert "Month Overview" in wb.sheetnames
        assert "Daily Breakdown" in wb.sheetnames
        assert "Group & Category Breakdown" in wb.sheetnames
        assert "Product-Wise Totals" in wb.sheetnames

        ws_grp = wb["Group & Category Breakdown"]
        headers = [ws_grp.cell(row=6, column=c).value for c in range(1, 7)]
        assert "Group" in headers
        assert "Category" in headers
        assert "Units Sold (Month)" in headers


def test_master_financial_group_and_category_breakdown(app, setup_catalog_and_bills):
    """Test Master Financial Sheet has dedicated Sales by Group & Category sheet."""
    with app.app_context():
        xlsx_svc = ExcelXLSXService()
        today = date.today()
        filepath = xlsx_svc.export_master_financial_sheet(today.year)

        wb = openpyxl.load_workbook(filepath)
        assert "Executive Summary" in wb.sheetnames
        assert "Sales Detail (Year)" in wb.sheetnames
        assert "Sales by Group & Category" in wb.sheetnames
        assert "Expense Detail (Year)" in wb.sheetnames

        ws_grp = wb["Sales by Group & Category"]
        headers = [ws_grp.cell(row=6, column=c).value for c in range(1, 8)]
        assert "Group" in headers
        assert "Category" in headers
        assert "Sales Revenue" in headers


def test_summary_service_group_and_category(app, setup_catalog_and_bills):
    """Test SummaryService returns group_totals and group_category_breakdown."""
    with app.app_context():
        db_svc = DatabaseService()
        summary_svc = SummaryService(db_svc)

        summary = summary_svc.get_today_summary()
        assert "group_totals" in summary
        assert "group_category_breakdown" in summary

        # Total sales is 280 (160 + 120)
        # Beverages: Masala Chai (40) + Mango Shake (120) = 160
        # Main Kitchen: Veg Cheese Burger (120) = 120
        assert summary["group_totals"]["Beverages"] == 160.0
        assert summary["group_totals"]["Main Kitchen"] == 120.0
        assert summary["group_category_breakdown"]["Beverages"]["Tea & Coffee"] == 40.0
        assert summary["group_category_breakdown"]["Beverages"]["Milkshakes"] == 120.0
        assert summary["group_category_breakdown"]["Main Kitchen"]["Burgers"] == 120.0


def test_csv_exports_group_and_category(app, setup_catalog_and_bills):
    """Test CSV export service includes Group and Category headers and data."""
    with app.app_context():
        db_svc = DatabaseService()
        summary_svc = SummaryService(db_svc)
        excel_svc = ExcelService()

        bills_data = [
            {
                "bill_no": 1001,
                "date": "2026-09-02",
                "time": "12:00:00",
                "total": 160.0,
                "products": [
                    {
                        "product_id": "P101",
                        "name": "Masala Chai",
                        "group": "Beverages",
                        "category": "Tea & Coffee",
                        "quantity": 2,
                        "price": 20.0,
                    }
                ],
            }
        ]

        # Test bills CSV export
        csv_path = excel_svc.export_today_sales_to_csv(bills_data)
        assert os.path.exists(csv_path)
        content = excel_svc.get_csv_content(csv_path)
        assert "Group" in content
        assert "Beverages" in content
        assert "Tea & Coffee" in content

        # Test summary CSV export
        sum_data = summary_svc.get_today_summary()
        sum_csv_path = excel_svc.export_summary_to_csv(sum_data)
        sum_content = excel_svc.get_csv_content(sum_csv_path)
        assert "=== GROUP WISE SALES ===" in sum_content
        assert "Beverages" in sum_content
